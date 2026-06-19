import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_import_calculator():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # Estilos de Diseño
    # -------------------------------------------------------------
    font_family = "Segoe UI"
    
    # Fuentes
    title_font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    section_text_font = Font(name=font_family, size=11, bold=True, color="C00000")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=10, bold=True, color="000000")
    regular_font = Font(name=font_family, size=10, color="000000")
    small_gray_font = Font(name=font_family, size=9, italic=True, color="595959")
    
    # Rellenos
    navy_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid") # Pizarra
    red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # Crimson
    section_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") # Rojo claro
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo suave
    zebra_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid") # Zebra
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Total gris
    accent_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Verde suave
    
    # Bordes
    thin_border_side = Side(style='thin', color='D3D3D3')
    thick_bottom_side = Side(style='double', color='000000')
    thin_top_side = Side(style='thin', color='000000')
    
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(top=thin_top_side, bottom=thick_bottom_side, left=thin_border_side, right=thin_border_side)
    
    # Alineación
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_title = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Formatos
    format_usd = "$#,##0.00"
    format_pen = "S/ #,##0.00"
    format_percent = "0.0%"
    format_weight = '#,##0.00" kg"'
    format_volume = '#,##0.00" CBM"'
    format_qty = "#,##0"

    # =============================================================
    # HOJA 3: Tarifario de Referencia
    # =============================================================
    ws3 = wb.active
    ws3.title = "3. Tarifario de Referencia"
    ws3.views.sheetView[0].showGridLines = True
    
    # Título Principal
    ws3.merge_cells("A1:C1")
    title_cell = ws3["A1"]
    title_cell.value = "TARIFARIO DE REFERENCIA Y VARIABLES ADM (MSI)"
    title_cell.font = title_font
    title_cell.fill = navy_fill
    title_cell.alignment = align_title
    ws3.row_dimensions[1].height = 40
    
    # Tarifas de origen y flete
    ws3.merge_cells("A3:C3")
    ws3["A3"] = "1. COSTOS DE ORIGEN Y FLETE"
    ws3["A3"].font = section_font
    ws3["A3"].fill = red_fill
    ws3["A3"].alignment = align_left
    ws3.row_dimensions[3].height = 24
    
    origen_flete = [
        ("Flete Marítimo por CBM (LCL)", 40.00, "Flete por metro cúbico"),
        ("BL Fee (USD)", 70.00, "Derecho de emisión de BL (origen)"),
        ("Pick Up (USD)", 80.00, "Recojo local en origen"),
        ("Gastos de Origen (USD)", 120.00, "Gastos locales del puerto de origen")
    ]
    
    r_idx = 4
    for label, val, note in origen_flete:
        ws3.cell(row=r_idx, column=1, value=label).font = regular_font
        ws3.cell(row=r_idx, column=1).alignment = align_left
        val_cell = ws3.cell(row=r_idx, column=2, value=val)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.number_format = format_usd
        val_cell.fill = input_fill
        ws3.cell(row=r_idx, column=3, value=note).font = small_gray_font
        for c in range(1, 4):
            ws3.cell(row=r_idx, column=c).border = border_cell
        ws3.row_dimensions[r_idx].height = 20
        r_idx += 1
        
    # Servicios Destino Callao
    r_idx += 1
    ws3.merge_cells(f"A{r_idx}:C{r_idx}")
    ws3[f"A{r_idx}"] = "2. SERVICIOS DESTINO CALLAO (NACIONALES)"
    ws3[f"A{r_idx}"].font = section_font
    ws3[f"A{r_idx}"].fill = red_fill
    ws3[f"A{r_idx}"].alignment = align_left
    ws3.row_dimensions[r_idx].height = 24
    r_idx += 1
    
    destino_params = [
        ("Seguro Comercial (USD)", 0.00, "Póliza local si aplica"),
        ("Almacenaje Verde LCL (USD)", 325.00, "Almacén temporal autorizado LCL green channel"),
        ("Transporte Interno (USD)", 0.00, "Flete nacional al almacén del cliente"),
        ("Comisión de Aduana (USD)", 80.00, "Honorarios de agente de aduana"),
        ("Gastos Operativos (USD)", 30.00, "Gastos operativos complementarios"),
        ("Visto Bueno LCL (USD)", 110.00, "Visto Bueno Callao consolidado"),
        ("Visto Bueno FCL (USD)", 0.00, "Naviera FCL"),
        ("Gate In FCL (USD)", 0.00, "Depósito contenedores FCL"),
        ("Descarga Puerto (USD)", 0.00, "Descarga física buque")
    ]
    
    for label, val, note in destino_params:
        ws3.cell(row=r_idx, column=1, value=label).font = regular_font
        ws3.cell(row=r_idx, column=1).alignment = align_left
        val_cell = ws3.cell(row=r_idx, column=2, value=val)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.number_format = format_usd
        val_cell.fill = input_fill
        ws3.cell(row=r_idx, column=3, value=note).font = small_gray_font
        for c in range(1, 4):
            ws3.cell(row=r_idx, column=c).border = border_cell
        ws3.row_dimensions[r_idx].height = 20
        r_idx += 1
        
    # Impuestos y Variables
    r_idx += 1
    ws3.merge_cells(f"A{r_idx}:C{r_idx}")
    ws3[f"A{r_idx}"] = "3. TASAS DE IMPUESTOS Y VARIABLES"
    ws3[f"A{r_idx}"].font = section_font
    ws3[f"A{r_idx}"].fill = red_fill
    ws3[f"A{r_idx}"].alignment = align_left
    ws3.row_dimensions[r_idx].height = 24
    r_idx += 1
    
    tax_params = [
        ("Arancel Ad-Valorem (%)", 0.00, "Tasa según partida arancelaria"),
        ("Percepción del IGV (%)", 0.035, "Tasa de percepción (3.5% default)"),
        ("Tipo de Cambio (S/.)", 3.78, "Tipo de cambio oficial (SUNAT)"),
        ("Doc. Fee (USD)", 30.00, "Gastos de documentación destino"),
        ("Descarga por TN (USD)", 30.00, "Tarifa de descarga de mercancía LCL"),
        ("Flete Aéreo por kg (USD)", 4.80, "Tarifa de flete aéreo por kilogramo"),
        ("Flete Marítimo FCL Flat (USD)", 1500.00, "Tarifa plana de flete por contenedor FCL")
    ]
    
    for label, val, note in tax_params:
        ws3.cell(row=r_idx, column=1, value=label).font = regular_font
        ws3.cell(row=r_idx, column=1).alignment = align_left
        val_cell = ws3.cell(row=r_idx, column=2, value=val)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.fill = input_fill
        
        if "%" in label or "Arancel" in label or "Percepción" in label:
            val_cell.number_format = format_percent
        elif "Tipo de Cambio" in label:
            val_cell.number_format = "0.000"
        else:
            val_cell.number_format = format_usd
            
        ws3.cell(row=r_idx, column=3, value=note).font = small_gray_font
        for c in range(1, 4):
            ws3.cell(row=r_idx, column=c).border = border_cell
        ws3.row_dimensions[r_idx].height = 20
        r_idx += 1
        
    ws3.column_dimensions['A'].width = 38
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 50

    # =============================================================
    # HOJA 1: Calculadora de Cliente
    # =============================================================
    ws1 = wb.create_sheet(title="1. Calculadora de Cliente", index=0)
    ws1.views.sheetView[0].showGridLines = True
    
    # Título Principal
    ws1.merge_cells("A1:D1")
    title_cell1 = ws1["A1"]
    title_cell1.value = "MSI CARGO - SIMULADOR COMERCIAL DE IMPORTACIÓN"
    title_cell1.font = title_font
    title_cell1.fill = navy_fill
    title_cell1.alignment = align_title
    ws1.row_dimensions[1].height = 40
    
    # Entradas de Datos
    ws1.merge_cells("A3:C3")
    ws1["A3"] = "DATOS DE LA MERCANCÍA"
    ws1["A3"].font = section_font
    ws1["A3"].fill = red_fill
    ws1["A3"].alignment = align_left
    ws1.row_dimensions[3].height = 24
    
    inputs_client = [
        ("Valor de la Mercadería (USD)", 3949.00, "Precio neto según factura comercial", format_usd),
        ("Peso Neto Total (kg)", 1000.00, "Peso neto físico total de la carga (Mín. 1000 kg para flete y descarga)", format_qty),
        ("Volumen Total (CBM)", 2.14, "Volumen en metros cúbicos total (Mín. 1 CBM para flete)", "0.00"),
        ("Incoterm", "EXW", "Seleccionar 'EXW', 'FOB' o 'CIF'", "@"),
        ("Tipo de Transporte", "Marítimo LCL", "Seleccionar 'Marítimo LCL' o 'Aéreo'", "@"),
        ("Origen de la Mercancía", "China", "Seleccionar origen de la carga (China, EEUU, Europa, Otros)", "@")
    ]
    
    r_idx = 4
    for label, val, note, fmt in inputs_client:
        ws1.cell(row=r_idx, column=1, value=label).font = regular_font
        ws1.cell(row=r_idx, column=1).alignment = align_left
        
        val_cell = ws1.cell(row=r_idx, column=2, value=val)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.fill = input_fill
        val_cell.number_format = fmt
        
        ws1.cell(row=r_idx, column=3, value=note).font = small_gray_font
        for c in range(1, 4):
            ws1.cell(row=r_idx, column=c).border = border_cell
        ws1.row_dimensions[r_idx].height = 20
        r_idx += 1
        
    # Data Validation para desplegables
    dv_incoterm = DataValidation(type="list", formula1='"EXW,FOB,CIF"', allow_blank=True)
    ws1.add_data_validation(dv_incoterm)
    dv_incoterm.add(ws1["B7"]) # Celda B7 es el Incoterm
    
    dv_transporte = DataValidation(type="list", formula1='"Marítimo LCL,Marítimo FCL,Aéreo"', allow_blank=True)
    ws1.add_data_validation(dv_transporte)
    dv_transporte.add(ws1["B8"]) # Celda B8 es el Tipo de Transporte
    
    dv_origen = DataValidation(type="list", formula1='"China,EEUU,Europa,Otros"', allow_blank=True)
    ws1.add_data_validation(dv_origen)
    dv_origen.add(ws1["B9"]) # Celda B9 es el Origen
    
    # Resumen de Resultados Destacados
    r_idx += 1
    ws1.merge_cells(f"A{r_idx}:C{r_idx}")
    ws1[f"A{r_idx}"] = "RESUMEN DE COTIZACIÓN COMERCIAL"
    ws1[f"A{r_idx}"].font = section_font
    ws1[f"A{r_idx}"].fill = red_fill
    ws1[f"A{r_idx}"].alignment = align_left
    ws1.row_dimensions[r_idx].height = 24
    r_idx += 1
    
    # Caja destacada para Costo Total (Row 12)
    total_row_usd = r_idx
    ws1.cell(row=r_idx, column=1, value="COSTO TOTAL DE IMPORTACIÓN (USD)").font = bold_font
    total_val_usd = ws1.cell(row=r_idx, column=2, value="='2. Estructura de Costeo (MSI)'!H63")
    total_val_usd.font = Font(name=font_family, size=11, bold=True, color="C00000")
    total_val_usd.number_format = format_usd
    total_val_usd.fill = accent_green_fill
    total_val_usd.alignment = align_right
    ws1.cell(row=r_idx, column=3, value="Fórmula: Mercadería + Impuestos + Servicios").font = small_gray_font
    for c in range(1, 4):
        ws1.cell(row=r_idx, column=c).border = border_cell
    ws1.row_dimensions[r_idx].height = 22
    r_idx += 1
    
    # Row 13
    total_row_pen = r_idx
    ws1.cell(row=r_idx, column=1, value="COSTO TOTAL DE IMPORTACIÓN (S/.)").font = bold_font
    total_val_pen = ws1.cell(row=r_idx, column=2, value=f"=B{total_row_usd}*'3. Tarifario de Referencia'!$B$23")
    total_val_pen.font = Font(name=font_family, size=11, bold=True, color="C00000")
    total_val_pen.number_format = format_pen
    total_val_pen.fill = accent_green_fill
    total_val_pen.alignment = align_right
    ws1.cell(row=r_idx, column=3, value="Convertido al tipo de cambio configurado").font = small_gray_font
    for c in range(1, 4):
        ws1.cell(row=r_idx, column=c).border = border_cell
    ws1.row_dimensions[r_idx].height = 22
    r_idx += 1
    
    # Detalle de los Tres Pilares
    r_idx += 1
    ws1.merge_cells(f"A{r_idx}:C{r_idx}")
    ws1[f"A{r_idx}"] = "DISTRIBUCIÓN DE PILARES DE COSTO"
    ws1[f"A{r_idx}"].font = section_text_font
    ws1[f"A{r_idx}"].alignment = align_left
    ws1.row_dimensions[r_idx].height = 20
    r_idx += 1
    
    breakdown_pilars = [
        ("Valor Mercadería (Proveedor China)", "='2. Estructura de Costeo (MSI)'!H60", "Costo ex-works de compra"),
        ("Impuestos de Importación (SUNAT)", "='2. Estructura de Costeo (MSI)'!H61", "Arancel + IGV/IPM + Percepción aduanera"),
        ("Servicios Logísticos (MSI Peru Cargo)", "='2. Estructura de Costeo (MSI)'!H62", "Flete + Origen (solo EXW) + Servicios Callao + IGV de servicios")
    ]
    
    for label, formula, note in breakdown_pilars:
        ws1.cell(row=r_idx, column=1, value=label).font = regular_font
        ws1.cell(row=r_idx, column=1).alignment = align_left
        val_cell = ws1.cell(row=r_idx, column=2, value=formula)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.number_format = format_usd
        ws1.cell(row=r_idx, column=3, value=note).font = small_gray_font
        for c in range(1, 4):
            ws1.cell(row=r_idx, column=c).border = border_cell
        ws1.row_dimensions[r_idx].height = 20
        r_idx += 1
        
    # Detalle de Liquidación Aduanera
    r_idx += 1
    ws1.merge_cells(f"A{r_idx}:C{r_idx}")
    ws1[f"A{r_idx}"] = "DESGLOSE DE LIQUIDACIÓN DE ADUANA"
    ws1[f"A{r_idx}"].font = section_text_font
    ws1[f"A{r_idx}"].alignment = align_left
    ws1.row_dimensions[r_idx].height = 20
    r_idx += 1
    
    liqui_items = [
        ("Valor FOB Aduanero", "='2. Estructura de Costeo (MSI)'!J8", "Base FOB incluyendo gastos de origen si es EXW"),
        ("Flete Aduanero", "='2. Estructura de Costeo (MSI)'!J9", "Prorrateado con mínimos y recargos (Cero si es CIF)"),
        ("Seguro Aduanero", "='2. Estructura de Costeo (MSI)'!J10", "Seguro presuntivo SUNAT (Cero si es CIF)"),
        ("Valor CIF (Base Imponible)", "='2. Estructura de Costeo (MSI)'!J11", "Valor en aduanas peruanas"),
        ("Derechos Arancelarios (Ad-Valorem)", "='2. Estructura de Costeo (MSI)'!J22", "Impuesto al arancel según partida"),
        ("IGV + IPM (18%)", "=('2. Estructura de Costeo (MSI)'!J26+'2. Estructura de Costeo (MSI)'!J27)", "IGV 16% e IPM 2% aduanero"),
        ("Percepción de Aduanas (SUNAT)", "='2. Estructura de Costeo (MSI)'!J32", "Pago adelantado de IGV (Habitual 3.5%)")
    ]
    
    for label, formula, note in liqui_items:
        is_cif = "CIF" in label
        ws1.cell(row=r_idx, column=1, value=label).font = bold_font if is_cif else regular_font
        ws1.cell(row=r_idx, column=1).alignment = align_left
        val_cell = ws1.cell(row=r_idx, column=2, value=formula)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.number_format = format_usd
        if is_cif:
            val_cell.fill = total_fill
        ws1.cell(row=r_idx, column=3, value=note).font = small_gray_font
        for c in range(1, 4):
            ws1.cell(row=r_idx, column=c).border = border_cell
        ws1.row_dimensions[r_idx].height = 20
        r_idx += 1
        
    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 50

    # =============================================================
    # HOJA 2: Estructura de Costeo (MSI)
    # =============================================================
    ws2 = wb.create_sheet(title="2. Estructura de Costeo (MSI)", index=1)
    ws2.views.sheetView[0].showGridLines = True
    
    # Título Principal
    ws2.merge_cells("A1:K1")
    title_cell2 = ws2["A1"]
    title_cell2.value = "ESTRUCTURA DE COSTEO DETALLADA - IMPORTACIÓN COMERCIAL"
    title_cell2.font = title_font
    title_cell2.fill = navy_fill
    title_cell2.alignment = align_title
    ws2.row_dimensions[1].height = 40
    
    # Metadata del Documento
    ws2["I2"] = "FECHA"
    ws2["J2"] = "2026-06-18"
    ws2["I3"] = "COTIZACIÓN N°"
    ws2["J3"] = "MSI-2026-LCL-01"
    ws2["I4"] = "CLIENTE ID"
    ws2["J4"] = "MSI-CLI-01"
    
    ws2["B6"] = "Asesor de ventas: Jesus Riojas"
    ws2["B7"] = "Transporte : Maritimo - EXW - Qingdao - LCL"
    ws2["B9"] = "Datos del importador:"
    ws2["B10"] = "Nombre: Raúl Bardales"
    ws2["B11"] = "Teléfono: 929 242 439"
    ws2["B12"] = "Regimen de Aduana : Importación para el consumo"
    
    ws2["B17"] = "Los montos en este documento son montos estimados, los mismos que podrían variar con los documentos finales."
    
    # Fila 7: Valor EXW
    ws2["I7"] = "VALOR EXW, FCA $"
    ws2["J7"] = "='1. Calculadora de Cliente'!B4"
    ws2["J7"].number_format = format_usd
    ws2["J7"].font = bold_font
    
    # Fila 8: Valor FOB (Formula depending on Incoterm, referencing B7 of Sheet 1)
    ws2["I8"] = "VALOR FOB $"
    ws2["J8"] = "=IF('1. Calculadora de Cliente'!$B$7=\"EXW\", J7+E42+E43+E44, J7)"
    ws2["J8"].number_format = format_usd
    ws2["J8"].font = bold_font
    
    # Fila 9: Flete (Formula with minimum CBM, weight TN, transport selector B8 and origin recargos B9)
    # B4: Maritimo CBM rate, B22: Aereo kg rate, B7: Incoterm, B8: Transporte, B9: Origen
    # cbm is B6, peso is B5
    ws2["I9"] = "FLETE"
    ws2["J9"] = "=IF('1. Calculadora de Cliente'!$B$7=\"CIF\", 0, IF('1. Calculadora de Cliente'!$B$8=\"Aéreo\", '1. Calculadora de Cliente'!$B$5*'3. Tarifario de Referencia'!$B$26, IF('1. Calculadora de Cliente'!$B$8=\"Marítimo FCL\", '3. Tarifario de Referencia'!$B$27, MAX(MAX('1. Calculadora de Cliente'!$B$6, 1.0), MAX('1. Calculadora de Cliente'!$B$5, 1000.0)/1000.0)*'3. Tarifario de Referencia'!$B$4)))*IF('1. Calculadora de Cliente'!$B$9=\"Europa\", 1.2, IF('1. Calculadora de Cliente'!$B$9=\"EEUU\", 1.1, 1.0)))"
    ws2["J9"].number_format = format_usd
    ws2["J9"].font = bold_font
    
    # Fila 10: Seguro
    ws2["I10"] = "SEGURO"
    ws2["J10"] = "=IF('1. Calculadora de Cliente'!$B$7=\"CIF\", 0, 1.5%*J8)"
    ws2["J10"].number_format = format_usd
    ws2["J10"].font = bold_font
    
    # Fila 11: Valor CIF
    ws2["I11"] = "VALOR CIF"
    ws2["J11"] = "=SUM(J8:J10)"
    ws2["J11"].number_format = format_usd
    ws2["J11"].font = bold_font
    
    # Fila 12: POL/POD
    ws2["I12"] = "POL/POD"
    ws2["J12"] = "Qingdao / Callao"
    
    # Fila 13: Peso
    ws2["I13"] = "PESO (TN) (KG)"
    ws2["J13"] = "='1. Calculadora de Cliente'!B5/1000"
    ws2["J13"].number_format = "0.000"
    
    # Fila 14: Volumen
    ws2["I14"] = "VOLUMEN (CBM) (PVO)"
    ws2["J14"] = "='1. Calculadora de Cliente'!B6"
    ws2["J14"].number_format = "0.00"
    
    ws2["I15"] = "FRECUENCIA"
    ws2["J15"] = "SEMANAL"
    ws2["I16"] = "T.T. APROX."
    ws2["J16"] = 30
    
    # Cabeceras de Impuestos (SUNAT)
    ws2["B21"] = "ITEM"
    ws2["E21"] = "DESCRIPCION"
    ws2["F21"] = "TASA %"
    ws2["H21"] = "BASE IMPONIBLE"
    ws2["J21"] = "DOLARES"
    
    for col in ["B", "E", "F", "H", "J"]:
        ws2[f"{col}21"].font = bold_font
        ws2[f"{col}21"].border = border_cell
        ws2[f"{col}21"].alignment = align_center
        
    # Fila 22: Ad-Valorem
    ws2["B22"] = "1.1"
    ws2["E22"] = "AD-VALOREM"
    ws2["F22"] = "='3. Tarifario de Referencia'!B21"
    ws2["F22"].number_format = format_percent
    ws2["H22"] = "VALOR CIF"
    ws2["J22"] = "=J11*F22"
    ws2["J22"].number_format = format_usd
    
    # Fila 23: Sobretasa
    ws2["B23"] = "1.2"
    ws2["E23"] = "SOBRE TASA ADICIONAL"
    ws2["F23"] = 0
    ws2["F23"].number_format = format_percent
    ws2["H23"] = "VALOR CIF"
    ws2["J23"] = 0
    ws2["J23"].number_format = format_usd
    
    # Fila 24: Derechos especificos
    ws2["B24"] = "1.3"
    ws2["E24"] = "DERECHOS ESPECIFICOS"
    ws2["F24"] = 0
    ws2["F24"].number_format = format_percent
    ws2["H24"] = "VALOR FOB"
    ws2["J24"] = 0
    ws2["J24"].number_format = format_usd
    
    # Fila 25: ISC
    ws2["B25"] = "1.4"
    ws2["E25"] = "ISC"
    ws2["F25"] = 0
    ws2["F25"].number_format = format_percent
    ws2["H25"] = "CIF+A/V+ ST+DE"
    ws2["J25"] = "=J11*F25"
    ws2["J25"].number_format = format_usd
    
    # Fila 26: IGV
    ws2["B26"] = "1.5"
    ws2["E26"] = "IGV"
    ws2["F26"] = 0.16
    ws2["F26"].number_format = format_percent
    ws2["H26"] = "CIF+A/V+ ST+DE"
    ws2["J26"] = "=(J11+J22)*F26"
    ws2["J26"].number_format = format_usd
    
    # Fila 27: IPM
    ws2["B27"] = "1.6"
    ws2["E27"] = "IPM"
    ws2["F27"] = 0.02
    ws2["F27"].number_format = format_percent
    ws2["H27"] = "CIF+A/V+ ST+DE"
    ws2["J27"] = "=(J11+J22)*F27"
    ws2["J27"].number_format = format_usd
    
    # Aplicar fuentes y bordes a impuestos
    for r in range(22, 28):
        for col in ["B", "E", "F", "H", "J"]:
            ws2[f"{col}{r}"].border = border_cell
            if col in ["B", "F", "H"]:
                ws2[f"{col}{r}"].alignment = align_center
            elif col == "J":
                ws2[f"{col}{r}"].alignment = align_right
                ws2[f"{col}{r}"].font = bold_font
            else:
                ws2[f"{col}{r}"].alignment = align_left
                
    # Fila 29: Total Derechos
    ws2["H29"] = "TOTAL DERECHOS DE ADUANA:"
    ws2["H29"].font = bold_font
    ws2["J29"] = "=SUM(J22:J27)"
    ws2["J29"].font = bold_font
    ws2["J29"].number_format = format_usd
    ws2["J29"].alignment = align_right
    ws2["J29"].border = border_cell
    
    # Fila 30: CIF en US
    ws2["H30"] = "VALOR CIF EN US$ :"
    ws2["H30"].font = bold_font
    ws2["J30"] = "=J11"
    ws2["J30"].font = bold_font
    ws2["J30"].number_format = format_usd
    ws2["J30"].alignment = align_right
    ws2["J30"].border = border_cell
    
    # Fila 31: CIF + Derechos
    ws2["H31"] = "TOTAL DERECHOS DE ADUANA + VALOR CIF EN US$ :"
    ws2["H31"].font = bold_font
    ws2["J31"] = "=J29+J30"
    ws2["J31"].font = bold_font
    ws2["J31"].number_format = format_usd
    ws2["J31"].alignment = align_right
    ws2["J31"].border = border_cell
    
    # Fila 32: Percepcion
    ws2["H32"] = "PERCEPCIÓN DEL IGV"
    ws2["H32"].font = bold_font
    ws2["I32"] = "='3. Tarifario de Referencia'!B22"
    ws2["I32"].font = bold_font
    ws2["I32"].number_format = format_percent
    ws2["I32"].alignment = align_center
    ws2["J32"] = "=J31*I32"
    ws2["J32"].font = bold_font
    ws2["J32"].number_format = format_usd
    ws2["J32"].alignment = align_right
    ws2["J32"].border = border_cell
    
    # Fila 34: Impuestos de aduana a pagar
    ws2.merge_cells("E34:H34")
    ws2["E34"] = "IMPORTE APROXIMADO A PAGAR - IMPUESTOS DE ADUANA :"
    ws2["E34"].font = Font(name=font_family, size=10, bold=True, color="C00000")
    ws2["E34"].alignment = align_left
    ws2["J34"] = "=J29+J32"
    ws2["J34"].font = Font(name=font_family, size=11, bold=True, color="C00000")
    ws2["J34"].number_format = format_usd
    ws2["J34"].alignment = align_right
    ws2["J34"].border = border_cell
    ws2["J34"].fill = accent_green_fill
    
    ws2["B35"] = "NOTA: EL MONTO DE LOS IMPUESTOS DEBE SER ABONADO A LA CUENTA DE LA AGENCIA DE ADUANA EN SOLES AL T.C DEL DIA SUNAT"
    ws2["B35"].font = small_gray_font
    
    # SERVICIOS MSI
    ws2.merge_cells("B37:J37")
    ws2["B37"] = "SERVICIOS (SE PAGAN A MSI ADUANAS)"
    ws2["B37"].font = section_font
    ws2["B37"].fill = red_fill
    ws2["B37"].alignment = align_left
    ws2.row_dimensions[37].height = 24
    
    ws2["B39"] = "SERVICIOS DE CARGA INTERNACIONAL"
    ws2["B39"].font = bold_font
    ws2["H39"] = "SERVICIOS INTEGRAL DE ADUANA"
    ws2["H39"].font = bold_font
    
    # Filas de detalle logistico
    ws2["B41"] = "Flete Internacional X tn/m3 :"
    ws2["E41"] = "=J9"
    ws2["E41"].number_format = format_usd
    ws2["E41"].alignment = align_right
    ws2.merge_cells("H41:I41")
    ws2["H41"] = "Transporte interno"
    ws2["J41"] = "='3. Tarifario de Referencia'!B12"
    ws2["J41"].number_format = format_usd
    ws2["J41"].alignment = align_right
    ws2["K41"] = "+IGV"
    
    ws2["B42"] = "BL Fee:"
    ws2["E42"] = "=IF('1. Calculadora de Cliente'!$B$7=\"EXW\", '3. Tarifario de Referencia'!B5, 0)"
    ws2["E42"].number_format = format_usd
    ws2["E42"].alignment = align_right
    ws2["F42"] = "80" # Nota visual
    ws2.merge_cells("H42:I42")
    ws2["H42"] = "Comisión de Aduana :"
    ws2["J42"] = "='3. Tarifario de Referencia'!B13"
    ws2["J42"].number_format = format_usd
    ws2["J42"].alignment = align_right
    ws2["K42"] = "+IGV"
    
    ws2["B43"] = "Pick Up:"
    ws2["E43"] = "=IF('1. Calculadora de Cliente'!$B$7=\"EXW\", '3. Tarifario de Referencia'!B6, 0)"
    ws2["E43"].number_format = format_usd
    ws2["E43"].alignment = align_right
    ws2.merge_cells("H43:I43")
    ws2["H43"] = "Gastos Operativos :"
    ws2["J43"] = "='3. Tarifario de Referencia'!B14"
    ws2["J43"].number_format = format_usd
    ws2["J43"].alignment = align_right
    ws2["K43"] = "+IGV"
    
    ws2["B44"] = "Gastos de origen:"
    ws2["E44"] = "=IF('1. Calculadora de Cliente'!$B$7=\"EXW\", '3. Tarifario de Referencia'!B7, 0)"
    ws2["E44"].number_format = format_usd
    ws2["E44"].alignment = align_right
    ws2["F44"] = "120" # Nota
    
    ws2["B45"] = "Seguro:"
    ws2["E45"] = "='3. Tarifario de Referencia'!B10"
    ws2["E45"].number_format = format_usd
    ws2["E45"].alignment = align_right
    ws2["F45"] = "+IGV"
    
    ws2.merge_cells("H45:I45")
    ws2["H45"] = "COSTOS FACTURADOS POR TERCEROS"
    ws2["H45"].font = bold_font
    
    ws2["B46"] = "Doc. Fee. :"
    ws2["E46"] = "=IF('1. Calculadora de Cliente'!$B$8=\"Aéreo\", 0, '3. Tarifario de Referencia'!B24)"
    ws2["E46"].number_format = format_usd
    ws2["E46"].alignment = align_right
    ws2["F46"] = "+IGV"
    
    ws2["B47"] = "Descarga x tn. :"
    ws2["E47"] = "=IF('1. Calculadora de Cliente'!$B$8=\"Marítimo LCL\", '3. Tarifario de Referencia'!B25*MAX(J13, 1.0), 0)"
    ws2["E47"].number_format = format_usd
    ws2["E47"].alignment = align_right
    ws2["F47"] = "+IGV"
    
    ws2.merge_cells("H47:I47")
    ws2["H47"] = "COSTO APROXIMADO DEL ALMACENAJE CANAL VERDE :"
    ws2["J47"] = "=IF('1. Calculadora de Cliente'!$B$8=\"Marítimo LCL\", '3. Tarifario de Referencia'!B11, 0)"
    ws2["J47"].number_format = format_usd
    ws2["J47"].alignment = align_right
    ws2["K47"] = "+IGV"
    
    ws2["B48"] = "Visto Bueno :"
    ws2["E48"] = "=IF('1. Calculadora de Cliente'!$B$8=\"Marítimo LCL\", '3. Tarifario de Referencia'!B15, 0)"
    ws2["E48"].number_format = format_usd
    ws2["E48"].alignment = align_right
    ws2["F48"] = "+IGV"
    
    ws2.merge_cells("H48:I48")
    ws2["H48"] = "COSTO APROXIMADO DEL VISTO BUENO DE LA LINEA NAVIERA (FCL)"
    ws2["J48"] = "=IF('1. Calculadora de Cliente'!$B$8=\"Marítimo FCL\", '3. Tarifario de Referencia'!B16, 0)"
    ws2["J48"].number_format = format_usd
    ws2["J48"].alignment = align_right
    ws2["K48"] = "+IGV"
    
    ws2.merge_cells("H49:I49")
    ws2["H49"] = "GATE IN APROX. DE LA LINEA FCL :"
    ws2["J49"] = "=IF('1. Calculadora de Cliente'!$B$8=\"Marítimo FCL\", '3. Tarifario de Referencia'!B17, 0)"
    ws2["J49"].number_format = format_usd
    ws2["J49"].alignment = align_right
    ws2["K49"] = "+IGV"
    
    ws2.merge_cells("H50:I50")
    ws2["H50"] = "COSTO APROXIMADO DE LA DESCARGA DE PUERTO :"
    ws2["J50"] = "=IF('1. Calculadora de Cliente'!$B$8=\"Marítimo FCL\", '3. Tarifario de Referencia'!B18, 0)"
    ws2["J50"].number_format = format_usd
    ws2["J50"].alignment = align_right
    ws2["K50"] = "+IGV"
    
    # Aplicar bordes delgados a las celdas de servicios
    for r in range(41, 51):
        for col in ["B", "E", "F", "H", "I", "J", "K"]:
            ws2[f"{col}{r}"].border = border_cell
            if col in ["E", "J"] and ws2[f"{col}{r}"].value is not None:
                ws2[f"{col}{r}"].font = bold_font
                    
    # Fila 52: Total Servicios
    ws2.merge_cells("H52:I52")
    ws2["H52"] = "TOTAL SERVICIOS"
    ws2["H52"].font = bold_font
    ws2["J52"] = "=SUM(E41:E48) + SUM(J41:J43) + SUM(J47:J50)"
    ws2["J52"].font = bold_font
    ws2["J52"].number_format = format_usd
    ws2["J52"].alignment = align_right
    ws2["J52"].border = border_cell
    
    # Fila 54: Total IGV de Servicios (grava solo servicios nacionales)
    ws2.merge_cells("H54:I54")
    ws2["H54"] = "TOTAL I.G.V."
    ws2["H54"].font = bold_font
    ws2["J54"] = "=(SUM(E45:E48) + SUM(J41:J43) + SUM(J47:J50))*18%"
    ws2["J54"].font = bold_font
    ws2["J54"].number_format = format_usd
    ws2["J54"].alignment = align_right
    ws2["J54"].border = border_cell
    
    # Fila 56: Total Proforma
    ws2.merge_cells("H56:I56")
    ws2["H56"] = "TOTAL PROFORMA"
    ws2["H56"].font = bold_font
    ws2["J56"] = "=SUM(J52:J54)"
    ws2["J56"].font = bold_font
    ws2["J56"].number_format = format_usd
    ws2["J56"].alignment = align_right
    ws2["J56"].border = border_cell
    ws2["J56"].fill = accent_green_fill
    
    # Resumen Total de Importacion
    ws2.merge_cells("E59:J59")
    ws2["E59"] = "RESUMEN TOTAL DE IMPORTACION"
    ws2["E59"].font = section_font
    ws2["E59"].fill = navy_fill
    ws2["E59"].alignment = align_left
    ws2.row_dimensions[59].height = 24
    
    ws2["E60"] = "VALOR DE LA MERCANCÍA"
    ws2["H60"] = "=J7"
    ws2["I60"] = "PROVEEDOR EN CHINA"
    
    ws2["E61"] = "IMPUESTOS"
    ws2["H61"] = "=J34"
    ws2["I61"] = "SUNAT"
    
    ws2["E62"] = "SERVICIOS LOGÍSTICOS"
    ws2["H62"] = "=J56"
    ws2["I62"] = "NOSOTROS"
    
    ws2["E63"] = "COSTO TOTAL DE IMPORTACIÓN"
    ws2["E63"].font = bold_font
    ws2["H63"] = "=SUM(H60:H62)"
    ws2["H63"].font = Font(name=font_family, size=11, bold=True, color="C00000")
    ws2["H63"].fill = accent_green_fill
    ws2["I63"] = "TOTAL GENERAL"
    ws2["I63"].font = bold_font
    
    for r in range(60, 64):
        for col in ["E", "H", "I"]:
            cell = ws2[f"{col}{r}"]
            cell.border = border_cell
            if col == "H":
                cell.alignment = align_right
                cell.number_format = format_usd
                if r != 63:
                    cell.font = bold_font
            else:
                cell.alignment = align_left
                
    ws2.column_dimensions['B'].width = 32
    ws2.column_dimensions['E'].width = 16
    ws2.column_dimensions['F'].width = 10
    ws2.column_dimensions['H'].width = 38
    ws2.column_dimensions['I'].width = 16
    ws2.column_dimensions['J'].width = 16
    ws2.column_dimensions['K'].width = 10

    # Guardar libro
    file_path = "Calculadora_Importacion.xlsx"
    wb.save(file_path)
    print(f"Libro de Excel creado exitosamente en: '{file_path}'")

if __name__ == "__main__":
    create_import_calculator()
